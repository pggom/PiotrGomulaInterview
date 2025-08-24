import sys, os, subprocess
import soundfile as sf
import numpy as np
import openpyxl
import pandas as pd
from tkinter import  ttk, messagebox
from openai import OpenAI


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PATH_TO_XLS = os.path.join(BASE_DIR, "Text_Files", "słówka_ang_dane.xlsx")
AUDIO_BASE = r"C:\Users\admin\Documents\ANGIELSKI\Angielski do słuchania"

if not os.path.exists(PATH_TO_XLS):
    error_message = ttk.Tk()
    error_message.withdraw()
    messagebox.showinfo("Błąd ścieżki", "Ścieżka nie istnieje")


def check_audio_file(english_word, file_path):


    if not os.path.isfile(file_path):
        return False

    elif not file_path.lower().endswith('.mp3'):
        return False
    else:
        min_size_kb = 10
        file_size_kb = os.path.getsize(file_path) / 1024
        if file_size_kb >= min_size_kb:

            return True
        else:
            print("Plik audio jest za mały.")
            return False

# def create_dict_from_xls(PATH_TO_XLS):
#
#
#     # Wczytanie danych z pliku xls
#     xls_data = pd.read_excel(PATH_TO_XLS)
#
#     # Inicjalizacja pustego słownika
#     not_used_pl_word_dict = {}
#     used_pl_word_dict = {}
#
#     # Iteracja po wierszach danych
#     for index, row in xls_data.iterrows():
#         audio_file_path = str(row.iloc[5])
#         row_number = index + 2
#
#
#         if audio_file_path == "nan":
#
#             # Pobranie słówka po polsku i angielsku
#             polish_word = row.iloc[1]
#
#             english_word = row.iloc[2]
#             not_used_pl_word_dict[polish_word] = row_number
#         else:
#             english_word = row.iloc[2]
#             if check_audio_file(english_word, audio_file_path) == False:
#                 print(f"Słowo {row.iloc[2]} ma nieprawidłowy plik audio.")
#             polish_word = row.iloc[1]
#             used_pl_word_dict[polish_word] = row_number
#     return not_used_pl_word_dict, used_pl_word_dict

def create_dict_from_xls(PATH_TO_XLS):
    xls_data = pd.read_excel(PATH_TO_XLS)

    not_used_pl_word_dict = {}
    used_pl_word_dict = {}

    for index, row in xls_data.iterrows():
        audio_file_path = str(row.iloc[5])
        points = row.iloc[6]  # kolumna G
        row_number = index + 2

        polish_word = str(row.iloc[1])  # <<< wymuszenie stringa
        english_word = row.iloc[2]

        if audio_file_path == "nan":
            not_used_pl_word_dict[polish_word] = (row_number, points)
        else:
            if check_audio_file(english_word, audio_file_path) == False:
                print(f"Słowo {english_word} ma nieprawidłowy plik audio.")
            used_pl_word_dict[polish_word] = (row_number, points)

    return not_used_pl_word_dict, used_pl_word_dict

def refresh_word_dicts_and_list():
    """Zwraca: dict_key, dict_used, pl_word_list (bez dotykania current_idx)."""
    dict_key, dict_used = create_dict_from_xls(PATH_TO_XLS)
    pl_list = sorted(
        dict_key.keys(),
        key=lambda w: (dict_key[w][1], w.lower())
    )
    return dict_key, dict_used, pl_list




def excel_save(row_nr, file_path,*args):


    workbook = openpyxl.load_workbook(PATH_TO_XLS)
    sheet = workbook.worksheets[0]


    sheet[f'A{row_nr}'] = 'ok'
    sheet[f'F{row_nr}'] = file_path

    if args:
        sheet[f'D{row_nr}'] = args[0] if len(args) > 0 else None
        sheet[f'E{row_nr}'] = args[1] if len(args) > 1 else None


    # Zapisz zmiany
    workbook.save(PATH_TO_XLS)

def update_excel_word(row_nr: int, new_pl: str, new_en: str):
    workbook = openpyxl.load_workbook(PATH_TO_XLS)
    sheet = workbook.worksheets[0]
    sheet[f'B{row_nr}'] = new_pl.strip()
    sheet[f'C{row_nr}'] = new_en.strip()
    workbook.save(PATH_TO_XLS)

def _to_stereo(x: np.ndarray) -> np.ndarray:
    """Z mono -> stereo (duplikacja kanału). Jeśli już stereo, zwraca bez zmian."""
    if x.ndim == 1:
        return np.stack([x, x], axis=1)
    return x

def _resample_linear(x: np.ndarray, orig_sr: int, target_sr: int) -> np.ndarray:
    """Prosty, szybki resampling liniowy (działa dla 1D lub 2D [N, C])."""
    if orig_sr == target_sr:
        return x
    # x shape: (N,) lub (N,2)
    n_samples = x.shape[0]
    duration = n_samples / float(orig_sr)
    new_n = int(round(duration * target_sr))
    # przygotuj oś czasu
    old_t = np.linspace(0.0, duration, num=n_samples, endpoint=False, dtype=np.float64)
    new_t = np.linspace(0.0, duration, num=new_n, endpoint=False, dtype=np.float64)
    if x.ndim == 1:
        return np.interp(new_t, old_t, x).astype(np.float32)
    else:
        # kanałami
        ch0 = np.interp(new_t, old_t, x[:, 0])
        ch1 = np.interp(new_t, old_t, x[:, 1])
        return np.stack([ch0, ch1], axis=1).astype(np.float32)

def combine_audio(list_word_and_folder):
    # będziemy składać wyłącznie stereo, float32
    combined_audio = np.zeros((0, 2), dtype=np.float32)
    sample_rate = None  # docelowy SR

    # 1) Dźwięk otwarcia
    start_effect_path = r"C:\Users\admin\Documents\PROGRAMOWANIE\nauka_angielskiego\Paper Slide - Sound Effect-cut.wav"
    if os.path.exists(start_effect_path):
        start_audio, start_sr = sf.read(start_effect_path, dtype='float32')
        start_audio = _to_stereo(start_audio)
        sample_rate = start_sr
        combined_audio = np.concatenate([combined_audio, start_audio], axis=0)

    # 2) Reszta elementów (słowo PL, EN, zdania...)
    for element in list_word_and_folder[1:]:
        path_to_file = element.audio_file_path
        audio, sr = sf.read(path_to_file, dtype='float32')
        audio = _to_stereo(audio)

        # ustaw SR, jeśli jeszcze nieustalony (brak efektu startowego)
        if sample_rate is None:
            sample_rate = sr

        # dopasuj SR, jeśli różny
        if sr != sample_rate:
            audio = _resample_linear(audio, sr, sample_rate)

        # pauza w próbkach (stereo => (N,2))
        pause_samples = int(element.pause_duration) * sample_rate
        if pause_samples < 0:
            pause_samples = 0
        pause = np.zeros((pause_samples, 2), dtype=np.float32)

        combined_audio = np.concatenate([combined_audio, audio, pause], axis=0)

    # 3) Zapis
    output_folder = r"C:\Users\admin\Documents\ANGIELSKI\Angielski do słuchania"
    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, list_word_and_folder[2].content + ".mp3")

    sf.write(output_file, combined_audio, samplerate=sample_rate)
    print(f"Tworzenie pliku zakończone: {output_file}")
    return output_file



def generat_sentence_from_gpt(pl_word: str, en_word: str, cefr: str = "A2", mode: str = "phrase"):
    """
    mode: 'phrase' => krótka kolokacja/zwrot po POLSKU
          'sentence' => krótkie zdanie po POLSKU
    """
    client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
    instruction = (
        "Zwróć JEDNĄ krótką frazę po polsku (bez kropki), która naturalnie oddaje znaczenie "
        f"angielskiego słowa «{en_word}», zgodne z polskim hasłem «{pl_word}». "
        f"Poziom języka: {cefr}. Bez dodatkowych komentarzy, tylko sama fraza."
        if mode == "phrase" else
        "Zwróć JEDNO krótkie zdanie po polsku, które naturalnie używa znaczenia "
        f"angielskiego słowa «{en_word}», zgodne z polskim hasłem «{pl_word}». "
        f"Poziom języka: {cefr}. Bez dodatkowych komentarzy, tylko zdanie."
    )

    # mały few-shot, żeby model trzymał format
    msgs = [
        {"role": "system", "content": "Jesteś asystentem tworzącym krótkie, naturalne polskie wyrażenia do nauki słownictwa."},
        {"role": "user", "content": "PL: okoliczności | EN: circumstances | mode: phrase | level: A2"},
        {"role": "assistant", "content": "w takich okolicznościach"},
        {"role": "user", "content": f"PL: {pl_word} | EN: {en_word} | mode: {mode} | level: {cefr}\n{instruction}"}
    ]

    resp = client.chat.completions.create(
        model="gpt-4.1-nano",
        messages=msgs,
        temperature=0.2,
    )
    return resp.choices[0].message.content.strip()



def add_points(row_nr, points_to_add=3):
    workbook = openpyxl.load_workbook(PATH_TO_XLS)
    sheet = workbook.worksheets[0]
    col_points = 7  # kolumna G = numer 7 (A=1, B=2...)
    current_value = sheet.cell(row=row_nr, column=col_points).value
    if current_value is None:
        current_value = 0
    sheet.cell(row=row_nr, column=col_points).value = current_value + points_to_add
    sheet.cell(row=row_nr, column=8).value = sheet.cell(row=row_nr, column=8).value + 1
    workbook.save(PATH_TO_XLS)

def restart_app():
    # odpal ten sam skrypt w nowym procesie
    subprocess.Popen([sys.executable] + sys.argv, close_fds=True)
    # bezpiecznie zakończ bieżący proces
    os._exit(0)

